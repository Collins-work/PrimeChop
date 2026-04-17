from openpyxl import load_workbook

path = 'primechop_audit.xlsx'
wb = load_workbook(path, read_only=True, data_only=True)
print('sheets', wb.sheetnames, flush=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print('sheet', sheet_name, 'rows', ws.max_row, 'cols', ws.max_column, flush=True)
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print('headers', headers, flush=True)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and any(isinstance(value, str) and 'WAI938' in value for value in row):
            print('MATCH', row, flush=True)
