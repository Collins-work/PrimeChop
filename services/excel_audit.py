from __future__ import annotations

import atexit
import logging
from pathlib import Path
from queue import Empty, Queue
import sqlite3
from threading import RLock
from threading import Event, Thread
import time
from typing import Iterable

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

try:
    import gspread  # type: ignore[import-untyped]
except ImportError:  # pragma: no cover - optional dependency
    gspread = None


logger = logging.getLogger(__name__)

ORDER_SHEET = "OrdersAudit"
WAITER_SHEET = "WaiterRegistry"

ORDER_HEADERS = [
    "event",
    "timestamp",
    "order_ref",
    "customer_id",
    "customer_name",
    "waiter_id",
    "waiter_name",
    "item",
    "amount",
    "hall",
    "room",
    "order_status",
    "payment_status",
    "payment_provider",
    "payment_tx_ref",
]

WAITER_HEADERS = [
    "user_id",
    "full_name",
    "waiter_code",
    "role",
    "verified",
    "online",
    "updated_at",
]


class ExcelAuditTrail:
    def __init__(
        self,
        file_path: str,
        enabled: bool = True,
        backend: str = "sqlite",
        sqlite_db_path: str = "primechop.db",
        google_spreadsheet_id: str = "",
        google_credentials_file: str = "",
        order_sheet_name: str = ORDER_SHEET,
        waiter_sheet_name: str = WAITER_SHEET,
        async_writes: bool = True,
        flush_interval_seconds: float = 1.0,
        max_batch_size: int = 25,
    ):
        self.enabled = enabled
        self.file_path = Path(file_path)
        self.backend = (backend or "sqlite").strip().lower()
        self.sqlite_db_path = Path(sqlite_db_path).expanduser()
        self.google_spreadsheet_id = (google_spreadsheet_id or "").strip()
        self.google_credentials_file = Path(google_credentials_file).expanduser() if google_credentials_file else None
        self.order_sheet_name = order_sheet_name or ORDER_SHEET
        self.waiter_sheet_name = waiter_sheet_name or WAITER_SHEET
        self.async_writes = bool(async_writes)
        self.flush_interval_seconds = max(0.2, float(flush_interval_seconds or 1.0))
        self.max_batch_size = max(1, int(max_batch_size or 25))
        self._google_client = None
        self._google_spreadsheet = None
        self._lock = RLock()
        self._queue: Queue[dict] = Queue()
        self._stop_event = Event()
        self._worker_thread: Thread | None = None

        if self.backend not in {"sqlite", "excel", "google"}:
            logger.warning("Unsupported audit backend '%s'. Falling back to sqlite.", self.backend)
            self.backend = "sqlite"

        if self.backend == "google":
            if gspread is None:
                logger.warning("Google Sheets backend requested but gspread is not installed. Falling back to sqlite.")
                self.backend = "sqlite"
            elif not self.google_spreadsheet_id or not self.google_credentials_file:
                logger.warning(
                    "Google Sheets backend requested but credentials/spreadsheet id missing. Falling back to sqlite."
                )
                self.backend = "sqlite"

        if not self.enabled:
            self.async_writes = False

        if self.enabled:
            self._ensure_storage()
            if self.async_writes:
                self._worker_thread = Thread(target=self._worker_loop, daemon=True, name="excel-audit-writer")
                self._worker_thread.start()
                atexit.register(self.close)

    def log_order(
        self,
        *,
        event: str,
        timestamp: str,
        order_ref: str,
        customer_id: int,
        customer_name: str,
        waiter_id: int,
        waiter_name: str,
        item: str,
        amount: int,
        hall: str,
        room: str,
        order_status: str,
        payment_status: str,
        payment_provider: str,
        payment_tx_ref: str,
    ) -> None:
        if not self.enabled:
            return
        values = [
            event,
            timestamp,
            order_ref,
            int(customer_id),
            customer_name,
            int(waiter_id),
            waiter_name,
            item,
            int(amount),
            hall,
            room,
            order_status,
            payment_status,
            payment_provider,
            payment_tx_ref,
        ]
        self._enqueue({"type": "order", "values": values})

    def upsert_waiter(
        self,
        *,
        user_id: int,
        full_name: str,
        waiter_code: str,
        role: str,
        verified: bool,
        online: bool,
        updated_at: str,
    ) -> None:
        if not self.enabled:
            return
        values = [
            int(user_id),
            full_name,
            waiter_code,
            role,
            1 if verified else 0,
            1 if online else 0,
            updated_at,
        ]
        self._enqueue({"type": "waiter_upsert", "values": values})

    def remove_waiter(self, user_id: int) -> None:
        if not self.enabled:
            return
        self._enqueue({"type": "waiter_remove", "user_id": int(user_id)})

    def sync_waiters(self, waiters: Iterable[dict]) -> None:
        if not self.enabled:
            return
        try:
            with self._lock:
                for waiter in waiters:
                    user_id = int(waiter.get("user_id", 0) or 0)
                    if user_id <= 0:
                        continue
                    self.upsert_waiter(
                        user_id=user_id,
                        full_name=waiter.get("full_name") or "",
                        waiter_code=waiter.get("waiter_code") or "",
                        role=waiter.get("role") or "customer",
                        verified=bool(int(waiter.get("waiter_verified") or 0)),
                        online=bool(int(waiter.get("waiter_online") or 0)),
                        updated_at=waiter.get("updated_at") or "",
                    )
        except Exception:
            logger.exception("Failed to sync waiter registry")

    def get_google_sheet_url(self) -> str:
        if self.backend != "google" or not self.google_spreadsheet_id:
            return ""
        return f"https://docs.google.com/spreadsheets/d/{self.google_spreadsheet_id}"

    def close(self) -> None:
        if not self.async_writes:
            return
        self._stop_event.set()
        if self._worker_thread and self._worker_thread.is_alive():
            self._worker_thread.join(timeout=self.flush_interval_seconds + 2.0)
        self._drain_queue_and_flush()

    def _ensure_storage(self) -> None:
        with self._lock:
            if self.backend == "google":
                self._ensure_google_sheet(self.order_sheet_name, ORDER_HEADERS)
                self._ensure_google_sheet(self.waiter_sheet_name, WAITER_HEADERS)
                return
            if self.backend == "sqlite":
                self._ensure_sqlite_tables()
                return
            workbook = self._load_or_create_workbook()
            self._ensure_sheet(workbook, self.order_sheet_name, ORDER_HEADERS)
            self._ensure_sheet(workbook, self.waiter_sheet_name, WAITER_HEADERS)
            self._save_workbook(workbook)

    def _sqlite_connection(self) -> sqlite3.Connection:
        self.sqlite_db_path.parent.mkdir(parents=True, exist_ok=True)
        connection = sqlite3.connect(self.sqlite_db_path)
        try:
            return connection
        except Exception:
            connection.close()
            raise

    def _ensure_sqlite_tables(self) -> None:
        with self._sqlite_connection() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS audit_orders (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    event TEXT NOT NULL,
                    timestamp TEXT NOT NULL,
                    order_ref TEXT NOT NULL,
                    customer_id INTEGER NOT NULL,
                    customer_name TEXT,
                    waiter_id INTEGER,
                    waiter_name TEXT,
                    item TEXT,
                    amount INTEGER NOT NULL,
                    hall TEXT,
                    room TEXT,
                    order_status TEXT,
                    payment_status TEXT,
                    payment_provider TEXT,
                    payment_tx_ref TEXT,
                    created_at TEXT NOT NULL DEFAULT (datetime('now'))
                )
                """
            )
            existing_columns = {
                row[1] for row in conn.execute("PRAGMA table_info(audit_orders)").fetchall()
            }
            if "waiter_id" not in existing_columns:
                conn.execute("ALTER TABLE audit_orders ADD COLUMN waiter_id INTEGER")
            if "waiter_name" not in existing_columns:
                conn.execute("ALTER TABLE audit_orders ADD COLUMN waiter_name TEXT")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_audit_orders_order_ref ON audit_orders(order_ref)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_audit_orders_timestamp ON audit_orders(timestamp)")

            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS audit_waiters (
                    user_id INTEGER PRIMARY KEY,
                    full_name TEXT,
                    waiter_code TEXT,
                    role TEXT,
                    verified INTEGER NOT NULL DEFAULT 0,
                    online INTEGER NOT NULL DEFAULT 0,
                    updated_at TEXT
                )
                """
            )

    def _load_or_create_workbook(self):
        if self.file_path.exists():
            return load_workbook(self.file_path)
        workbook = Workbook()
        default_sheet = workbook.active
        if default_sheet and default_sheet.title == "Sheet":
            workbook.remove(default_sheet)
        self._ensure_sheet(workbook, self.order_sheet_name, ORDER_HEADERS)
        self._ensure_sheet(workbook, self.waiter_sheet_name, WAITER_HEADERS)
        return workbook

    def _ensure_sheet(self, workbook, title: str, headers: list[str]):
        if title in workbook.sheetnames:
            sheet = workbook[title]
        else:
            sheet = workbook.create_sheet(title=title)

        header_missing = False
        for idx, header in enumerate(headers, start=1):
            if sheet.cell(row=1, column=idx).value != header:
                header_missing = True
                break

        if header_missing:
            for idx, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=idx, value=header)
        return sheet

    def _find_row_by_user_id(self, sheet, user_id: int) -> int | None:
        for row in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row, column=1).value
            try:
                if value is not None and int(value) == user_id:
                    return row
            except (TypeError, ValueError):
                continue
        return None

    def _save_workbook(self, workbook) -> None:
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(self.file_path)

    def _enqueue(self, operation: dict) -> None:
        if not self.enabled:
            return
        if not self.async_writes:
            self._process_batch([operation])
            return
        self._queue.put(operation)

    def _worker_loop(self) -> None:
        while not self._stop_event.is_set():
            batch = self._get_batch(wait_for_first=True)
            if not batch:
                continue
            self._process_batch(batch)
        self._drain_queue_and_flush()

    def _get_batch(self, wait_for_first: bool) -> list[dict]:
        first_timeout = self.flush_interval_seconds if wait_for_first else 0.0
        batch: list[dict] = []
        try:
            if first_timeout > 0:
                batch.append(self._queue.get(timeout=first_timeout))
            else:
                batch.append(self._queue.get_nowait())
        except Empty:
            return batch

        deadline = time.monotonic() + self.flush_interval_seconds
        while len(batch) < self.max_batch_size:
            remaining = deadline - time.monotonic()
            if remaining <= 0:
                break
            try:
                batch.append(self._queue.get(timeout=remaining))
            except Empty:
                break
        return batch

    def _drain_queue_and_flush(self) -> None:
        pending: list[dict] = []
        while True:
            try:
                pending.append(self._queue.get_nowait())
            except Empty:
                break
        if not pending:
            return
        for index in range(0, len(pending), self.max_batch_size):
            self._process_batch(pending[index : index + self.max_batch_size])

    def _process_batch(self, operations: list[dict]) -> None:
        if not operations:
            return
        try:
            with self._lock:
                if self.backend == "google":
                    self._process_google_batch(operations)
                elif self.backend == "sqlite":
                    self._process_sqlite_batch(operations)
                else:
                    self._process_excel_batch(operations)
        except Exception:
            logger.exception("Failed to flush audit batch")

    def _process_sqlite_batch(self, operations: list[dict]) -> None:
        order_rows: list[tuple] = []
        waiter_upserts: list[tuple] = []
        waiter_deletes: list[int] = []

        for operation in operations:
            op_type = operation.get("type")
            if op_type == "order":
                values = operation.get("values", [])
                if len(values) != len(ORDER_HEADERS):
                    continue
                order_rows.append(tuple(values))
            elif op_type == "waiter_upsert":
                values = operation.get("values", [])
                if len(values) != len(WAITER_HEADERS):
                    continue
                waiter_upserts.append(tuple(values))
            elif op_type == "waiter_remove":
                user_id = int(operation.get("user_id", 0) or 0)
                if user_id > 0:
                    waiter_deletes.append(user_id)

        if not order_rows and not waiter_upserts and not waiter_deletes:
            return

        with self._sqlite_connection() as conn:
            if order_rows:
                conn.executemany(
                    """
                    INSERT INTO audit_orders (
                        event,
                        timestamp,
                        order_ref,
                        customer_id,
                        customer_name,
                        waiter_id,
                        waiter_name,
                        item,
                        amount,
                        hall,
                        room,
                        order_status,
                        payment_status,
                        payment_provider,
                        payment_tx_ref
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    order_rows,
                )

            if waiter_upserts:
                conn.executemany(
                    """
                    INSERT INTO audit_waiters (
                        user_id,
                        full_name,
                        waiter_code,
                        role,
                        verified,
                        online,
                        updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(user_id) DO UPDATE SET
                        full_name=excluded.full_name,
                        waiter_code=excluded.waiter_code,
                        role=excluded.role,
                        verified=excluded.verified,
                        online=excluded.online,
                        updated_at=excluded.updated_at
                    """,
                    waiter_upserts,
                )

            if waiter_deletes:
                conn.executemany(
                    "DELETE FROM audit_waiters WHERE user_id=?",
                    [(user_id,) for user_id in waiter_deletes],
                )

    def _process_excel_batch(self, operations: list[dict]) -> None:
        workbook = self._load_or_create_workbook()
        order_sheet = self._ensure_sheet(workbook, self.order_sheet_name, ORDER_HEADERS)
        waiter_sheet = self._ensure_sheet(workbook, self.waiter_sheet_name, WAITER_HEADERS)

        for operation in operations:
            op_type = operation.get("type")
            if op_type == "order":
                order_sheet.append(operation.get("values", []))
            elif op_type == "waiter_upsert":
                values = operation.get("values", [])
                user_id = int(values[0]) if values else 0
                if user_id <= 0:
                    continue
                target_row = self._find_row_by_user_id(waiter_sheet, user_id)
                if target_row is None:
                    waiter_sheet.append(values)
                else:
                    for idx, value in enumerate(values, start=1):
                        waiter_sheet.cell(row=target_row, column=idx, value=value)
            elif op_type == "waiter_remove":
                user_id = int(operation.get("user_id", 0) or 0)
                if user_id <= 0:
                    continue
                target_row = self._find_row_by_user_id(waiter_sheet, user_id)
                if target_row is not None:
                    waiter_sheet.delete_rows(target_row, 1)

        self._save_workbook(workbook)

    def _process_google_batch(self, operations: list[dict]) -> None:
        order_rows: list[list] = []
        waiter_ops: list[dict] = []
        for operation in operations:
            if operation.get("type") == "order":
                order_rows.append(operation.get("values", []))
            else:
                waiter_ops.append(operation)

        if order_rows:
            worksheet = self._ensure_google_sheet(self.order_sheet_name, ORDER_HEADERS)
            self._append_google_rows(worksheet, order_rows)

        if not waiter_ops:
            return

        waiter_sheet = self._ensure_google_sheet(self.waiter_sheet_name, WAITER_HEADERS)
        for operation in waiter_ops:
            op_type = operation.get("type")
            if op_type == "waiter_upsert":
                self._upsert_google_waiter(operation.get("values", []), worksheet=waiter_sheet)
            elif op_type == "waiter_remove":
                user_id = int(operation.get("user_id", 0) or 0)
                if user_id <= 0:
                    continue
                target_row = self._find_google_row_by_user_id(waiter_sheet, user_id)
                if target_row is not None:
                    waiter_sheet.delete_rows(target_row)

    def _get_google_spreadsheet(self):
        if self._google_spreadsheet is not None:
            return self._google_spreadsheet
        if gspread is None:
            raise RuntimeError("gspread is not installed")
        if not self.google_credentials_file:
            raise RuntimeError("GOOGLE_SHEETS_CREDENTIALS_FILE is not configured")
        if not self.google_spreadsheet_id:
            raise RuntimeError("GOOGLE_SHEETS_SPREADSHEET_ID is not configured")
        self._google_client = gspread.service_account(filename=str(self.google_credentials_file))
        self._google_spreadsheet = self._google_client.open_by_key(self.google_spreadsheet_id)
        return self._google_spreadsheet

    def _ensure_google_sheet(self, title: str, headers: list[str]):
        spreadsheet = self._get_google_spreadsheet()
        try:
            worksheet = spreadsheet.worksheet(title)
        except Exception:
            worksheet = spreadsheet.add_worksheet(title=title, rows=1000, cols=max(len(headers), 20))

        first_row = worksheet.row_values(1)
        if first_row[: len(headers)] != headers:
            worksheet.update("A1", [headers], value_input_option="RAW")
        return worksheet

    def _append_google_rows(self, worksheet, rows: list[list]):
        if not rows:
            return
        if hasattr(worksheet, "append_rows"):
            worksheet.append_rows(rows, value_input_option="RAW")
            return
        for row in rows:
            worksheet.append_row(row, value_input_option="RAW")

    def _find_google_row_by_user_id(self, worksheet, user_id: int) -> int | None:
        values = worksheet.col_values(1)
        for row_index, value in enumerate(values[1:], start=2):
            try:
                if value and int(value) == user_id:
                    return row_index
            except (TypeError, ValueError):
                continue
        return None

    def _upsert_google_waiter(self, values: list, worksheet=None):
        worksheet = worksheet or self._ensure_google_sheet(self.waiter_sheet_name, WAITER_HEADERS)
        target_row = self._find_google_row_by_user_id(worksheet, int(values[0]))
        if target_row is None:
            worksheet.append_row(values, value_input_option="RAW")
            return
        end_col = chr(ord("A") + len(values) - 1)
        worksheet.update(f"A{target_row}:{end_col}{target_row}", [values], value_input_option="RAW")
